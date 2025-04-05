import time

import openpyxl
from selenium.webdriver.support.select import Select
from SeleniumProject.Data_driven_testing import XLUtil

from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.chrome.service import Service

serv_object = Service(r'C:\Users\Administrator\Desktop\PYTHON\Selenium\Driver\Chrome\chromedriver-win64\chromedriver.exe')
options = webdriver.ChromeOptions()
options.add_experimental_option('detach', True)

driver = webdriver.Chrome(service=serv_object, options=options)
driver.get("https://www.moneycontrol.com/fixed-income/calculator/state-bank-of-india-sbi/fixed-deposit-calculator-SBI-BSB001.html?classic=true")
driver.maximize_window()
driver.implicitly_wait(10)

driver.find_element(By.XPATH, "//*[@id='wzrk-confirm']").click()
file_name = "sample.xlsx"
wb = openpyxl.load_workbook(file_name)
ws = wb.active

max_r = XLUtil.getRowCount(file_name, ws.title)
max_c = XLUtil.getColumnCount(file_name, ws.title)

for r in range(2, max_r +1):
    driver.find_element(By.XPATH, "//*[@id='principal']").send_keys( ws.cell(r, 1).value )
    driver.find_element(By.XPATH, "//*[@id='interest']").send_keys(  ws.cell(r, 2).value )
    driver.find_element(By.XPATH, "//*[@id='tenure']").send_keys( ws.cell(r, 3).value  )

    interest_type = Select(driver.find_element(By.XPATH, "//*[@id='frequency']"))
    interest_type.select_by_visible_text(  ws.cell(r, 4).value )

    driver.find_element(By.XPATH, "//*[@id='fdMatVal']/div[2]/a[1]").click()

    final_value = driver.find_element(By.XPATH, "//*[@id='resp_matval']/strong").text
    ws.cell(r,6).value = int(float(final_value))

    driver.find_element(By.XPATH, "//*[@id='fdMatVal']/div[2]/a[2]").click()
wb.save(file_name)
driver.close()